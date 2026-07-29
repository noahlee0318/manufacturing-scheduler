"""Validated Excel import for the manufacturing shift scheduler.

The public application accepts the same five-sheet workbook contract as the
original browser importer.  This module deliberately returns plain Python
dictionaries using the existing schema-v2 camelCase field names so the rest of
the scheduler can consume imported and bundled JSON data interchangeably.
"""

from __future__ import annotations

import math
import re
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook


MAX_WORKBOOK_BYTES = 10 * 1024 * 1024
MAX_ROUTES = 5_000
MAX_OPERATORS = 1_000
MAX_DEPENDENCIES = 10_000

REQUIRED_SHEETS: dict[str, str] = {
    "Approved Routes": "Approved Complete-Part Route Master",
    "Operator Skills": "Operator Qualification Matrix",
    "Attendance": "Daily Attendance and Shift Selection",
    "Parent-Child": "Approved Parent-Child Dependencies",
    "Settings": "Complete-Shift Scheduler Settings",
}

SKILL_NAMES: tuple[str, ...] = (
    "Spray Bars",
    "Tube",
    "HP Mid",
    "IP Mid",
    "HP Top",
    "IP Top",
    "LP's",
    "BVD Sub",
    "BVD Top",
    "T's",
    "ATB's",
    "LP Subs",
    "HP/IP Top",
    "HP/IP Subassembly",
)

ROUTE_HEADERS: tuple[str, ...] = (
    "Part",
    "Numeric Seq",
    "Seq Override",
    "Part + Operation",
    "Operation",
    "Work Center",
    "Routing Description",
    "Orange Bucket",
    "Suggested Skill",
    "Skill Override",
    "Exact P75 Hr/Pc",
    "P75 Override",
    "Move Min Override",
    "Samples",
    "History Status",
    "WC Source",
    "Orange Cell",
    "Notes",
)

OPERATOR_HEADERS: tuple[str, ...] = (
    "Employee ID",
    "Employee Name",
    "Source Shift",
    *SKILL_NAMES,
    "Data Quality Note",
)

ATTENDANCE_HEADERS: tuple[str, ...] = (
    "Employee ID",
    "Employee Name",
    "Scheduled Today?",
    "Present Today?",
    "Shift Today",
    "Hours Override",
    "Source Hours",
    "Source Days",
    "Data Quality Note",
)

DEPENDENCY_HEADERS: tuple[str, ...] = ("Parent Part", "Child Part")

SETTINGS_HEADERS: tuple[str, ...] = ("Setting", "Value", "Why it matters")

SHIFT_HEADERS: tuple[str, ...] = (
    "Shift",
    "Start",
    "End",
    "Available Hours",
    "Source / action",
)

REQUIRED_HEADERS: Mapping[str, Any] = {
    "Approved Routes": ROUTE_HEADERS,
    "Operator Skills": OPERATOR_HEADERS,
    "Attendance": ATTENDANCE_HEADERS,
    "Parent-Child": DEPENDENCY_HEADERS,
    "Settings": {
        "settings": SETTINGS_HEADERS,
        "shifts": SHIFT_HEADERS,
    },
}

SCHEDULING_RULES: dict[str, str] = {
    "Job input": "One complete part and total positive whole-number quantity per row",
    "Route": "All exact orange-approved operations, sorted by effective operation sequence",
    "Time": "Quantity × exact part-operation 75th-percentile hr/piece",
    "Dependency": (
        "When both are entered, every child unit and generated sub-batch must "
        "finish before the parent's first operation begins"
    ),
    "Concurrency": (
        "Different work centers can run simultaneously with different "
        "qualified employees"
    ),
    "Work center": "One sub-batch operation at a time per exact work center",
    "Operator": (
        "One sub-batch operation at a time; present and skill level at least "
        "the threshold"
    ),
    "Batch": (
        "Each whole unit may advance after its own preceding route operation "
        "finishes; it does not wait for the remaining quantity"
    ),
    "Cutoff": (
        "Each sub-batch operation must fit inside one contiguous employee "
        "availability window"
    ),
    "Dispatch": (
        "Active children before parents; then priority, due time, shorter route "
        "time, stable job order, unit, and route step"
    ),
}


class WorkbookValidationError(ValueError):
    """Raised when an uploaded workbook does not satisfy the import contract."""


def validate_workbook_file(file_name: str, size_bytes: int) -> None:
    """Validate the upload-level file name and 10 MiB size limit.

    Workbook contents are validated by :func:`parse_workbook_bytes`.
    """

    if not isinstance(file_name, str) or Path(file_name).suffix.lower() != ".xlsx":
        raise WorkbookValidationError("The scheduler import must be an .xlsx workbook.")
    if isinstance(size_bytes, bool) or not isinstance(size_bytes, int) or size_bytes < 0:
        raise WorkbookValidationError("Workbook size must be a non-negative whole number of bytes.")
    if size_bytes > MAX_WORKBOOK_BYTES:
        raise WorkbookValidationError(
            "The scheduler workbook must be 10 MiB or smaller."
        )


validate_workbook_upload = validate_workbook_file


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and math.isfinite(value) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalized(value: Any) -> str:
    return re.sub(r"\s+", " ", _text(value)).lower()


def _coerce_number(value: Any) -> float | int:
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)) and not isinstance(value, (date, datetime)):
        number: float | int = value
    elif isinstance(value, str):
        clean = value.strip()
        if not clean:
            raise ValueError
        try:
            number = float(clean)
        except ValueError as exc:
            raise ValueError from exc
    else:
        raise ValueError
    if not math.isfinite(float(number)):
        raise ValueError
    return number


def _optional_number(value: Any, label: str) -> float | int | None:
    if value is None or _text(value) == "":
        return None
    try:
        return _coerce_number(value)
    except (TypeError, ValueError, OverflowError) as exc:
        raise WorkbookValidationError(f"{label} must be a number or blank.") from exc


def _require_number(value: Any, label: str) -> float | int:
    if value is None or _text(value) == "":
        raise WorkbookValidationError(f"{label} must be a number.")
    try:
        return _coerce_number(value)
    except (TypeError, ValueError, OverflowError) as exc:
        raise WorkbookValidationError(f"{label} must be a number.") from exc


def _require_integer(
    value: Any,
    label: str,
    minimum: int,
    maximum: int,
) -> int:
    number = _require_number(value, label)
    if (
        float(number).is_integer()
        and minimum <= float(number) <= maximum
    ):
        return int(number)
    raise WorkbookValidationError(
        f"{label} must be a whole number from {minimum} to {maximum}."
    )


def _optional_integer(
    value: Any,
    label: str,
    minimum: int,
    maximum: int,
) -> int | None:
    number = _optional_number(value, label)
    if number is None:
        return None
    if (
        float(number).is_integer()
        and minimum <= float(number) <= maximum
    ):
        return int(number)
    raise WorkbookValidationError(
        f"{label} must be a whole number from {minimum} to {maximum}, or blank."
    )


def _yes_no(value: Any, fallback: bool, label: str) -> bool:
    clean = _normalized(value)
    if not clean:
        return fallback
    if clean in {"yes", "y", "true", "1"}:
        return True
    if clean in {"no", "n", "false", "0"}:
        return False
    raise WorkbookValidationError(f"{label} must be Yes, No, or blank.")


def _rows_for(workbook: Any, sheet_name: str) -> list[tuple[Any, ...]]:
    if sheet_name not in getattr(workbook, "sheetnames", ()):
        raise WorkbookValidationError(
            f'Required worksheet "{sheet_name}" is missing.'
        )
    worksheet = workbook[sheet_name]
    return [tuple(row) for row in worksheet.iter_rows(values_only=True)]


def _verify_title(
    rows: Sequence[Sequence[Any]],
    sheet_name: str,
    expected_title: str,
) -> None:
    title = _text(rows[0][0]) if rows and rows[0] else ""
    if title != expected_title:
        raise WorkbookValidationError(
            f'"{sheet_name}" is not recognized. Cell A1 must be '
            f'"{expected_title}".'
        )


def _find_header(
    rows: Sequence[Sequence[Any]],
    sheet_name: str,
    required_headers: Sequence[str],
    search_limit: int = 15,
) -> tuple[int, tuple[str, ...]]:
    for row_index, row in enumerate(rows[:search_limit]):
        headers = tuple(_text(value) for value in row)
        if all(header in headers for header in required_headers):
            return row_index, headers
    raise WorkbookValidationError(
        f'"{sheet_name}" is missing required columns: '
        f'{", ".join(required_headers)}.'
    )


def _cell(row: Sequence[Any], headers: Sequence[str], name: str) -> Any:
    try:
        index = headers.index(name)
    except ValueError:
        return None
    return row[index] if index < len(row) else None


_CLOCK_PATTERN = re.compile(
    r"^(\d{1,2}):(\d{2})(?::\d{2})?\s*(AM|PM)?$",
    re.IGNORECASE,
)


def _parse_clock(value: Any, label: str) -> int:
    if isinstance(value, datetime):
        return value.hour * 60 + value.minute
    if isinstance(value, time):
        return value.hour * 60 + value.minute
    if isinstance(value, timedelta):
        fraction = (value.total_seconds() / 86_400) % 1
        return round(fraction * 1_440) % 1_440
    if (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and math.isfinite(float(value))
    ):
        fraction = float(value) % 1
        return round(fraction * 1_440) % 1_440

    clean = _text(value)
    match = _CLOCK_PATTERN.fullmatch(clean)
    if not match:
        raise WorkbookValidationError(f"{label} must be a valid Excel time.")
    hour = int(match.group(1))
    minute = int(match.group(2))
    suffix = (match.group(3) or "").upper()
    maximum_hour = 12 if suffix else 23
    if minute > 59 or hour > maximum_hour or hour < 0:
        raise WorkbookValidationError(f"{label} must be a valid Excel time.")
    if suffix == "AM":
        hour = 0 if hour == 12 else hour
    elif suffix == "PM":
        hour = 12 if hour == 12 else hour + 12
    return hour * 60 + minute


def _clock_text(minutes: int | float) -> str:
    normalized_minutes = round(minutes) % 1_440
    return f"{normalized_minutes // 60:02d}:{normalized_minutes % 60:02d}"


def _parse_settings(
    rows: Sequence[Sequence[Any]],
    warnings: list[str],
) -> dict[str, Any]:
    settings_row, settings_headers = _find_header(
        rows,
        "Settings",
        SETTINGS_HEADERS,
        10,
    )
    settings_by_name: dict[str, Any] = {}
    for row in rows[settings_row + 1 :]:
        label = _text(_cell(row, settings_headers, "Setting"))
        if label:
            settings_by_name[label] = _cell(row, settings_headers, "Value")

    minimum_solo_skill_level = _require_integer(
        settings_by_name.get("Minimum solo skill level"),
        "Settings > Minimum solo skill level",
        1,
        3,
    )
    default_move_minutes = _require_number(
        settings_by_name.get("Default move minutes"),
        "Settings > Default move minutes",
    )
    if not 0 <= float(default_move_minutes) <= 240:
        raise WorkbookValidationError(
            "Settings > Default move minutes must be from 0 to 240."
        )
    work_center_capacity = _require_integer(
        settings_by_name.get("Work-center capacity"),
        "Settings > Work-center capacity",
        1,
        99,
    )
    if work_center_capacity != 1:
        raise WorkbookValidationError(
            "This application currently requires Work-center capacity to remain 1."
        )

    shift_row, shift_headers = _find_header(
        rows,
        "Settings",
        SHIFT_HEADERS,
        20,
    )
    shifts: list[dict[str, Any]] = []
    shift_ids: set[int] = set()
    for row_index in range(shift_row + 1, len(rows)):
        row = rows[row_index]
        raw_id = _cell(row, shift_headers, "Shift")
        if _text(raw_id) == "":
            continue
        try:
            _coerce_number(raw_id)
        except (TypeError, ValueError, OverflowError):
            if shifts:
                break
        shift_id = _require_integer(
            raw_id,
            f"Settings shift row {row_index + 1}",
            1,
            999,
        )
        if shift_id in shift_ids:
            raise WorkbookValidationError(
                f"Settings contains duplicate Shift {shift_id}."
            )
        shift_ids.add(shift_id)

        start_minutes = _parse_clock(
            _cell(row, shift_headers, "Start"),
            f"Shift {shift_id} start",
        )
        end_minutes = _parse_clock(
            _cell(row, shift_headers, "End"),
            f"Shift {shift_id} end",
        )
        if start_minutes == end_minutes:
            raise WorkbookValidationError(
                f"Shift {shift_id} start and end cannot match."
            )
        duration_minutes = (end_minutes - start_minutes + 1_440) % 1_440
        stated_hours = _optional_number(
            _cell(row, shift_headers, "Available Hours"),
            f"Settings shift row {row_index + 1} Available Hours",
        )
        if (
            stated_hours is not None
            and abs(float(stated_hours) * 60 - duration_minutes) > 1
        ):
            warnings.append(
                f"Shift {shift_id} Available Hours did not match its clock "
                "times; the clock times were used."
            )
        source = _text(_cell(row, shift_headers, "Source / action"))
        shifts.append(
            {
                "id": shift_id,
                "startMinutes": start_minutes,
                "endMinutes": end_minutes,
                "startTime": _clock_text(start_minutes),
                "endTime": _clock_text(end_minutes),
                "availableHours": duration_minutes / 60,
                "crossesMidnight": end_minutes <= start_minutes,
                "source": source,
                "placeholder": bool(re.search("placeholder", source, re.IGNORECASE)),
            }
        )

    if not shifts:
        raise WorkbookValidationError(
            "Settings does not contain any usable shifts."
        )
    shifts.sort(key=lambda shift: shift["id"])

    return {
        "minimumSoloSkillLevel": minimum_solo_skill_level,
        "defaultMoveMinutes": default_move_minutes,
        "workCenterCapacity": work_center_capacity,
        "hardShiftCutoff": True,
        "defaultBreaks": {
            "1": {
                "label": "Lunch",
                "startTime": "11:10",
                "durationMinutes": 40,
            },
            "2": {
                "label": "Dinner",
                "startTime": "20:55",
                "durationMinutes": 40,
            },
        },
        "shifts": shifts,
        "schedulingRules": dict(SCHEDULING_RULES),
    }


def _natural_key(value: str) -> tuple[Any, ...]:
    return tuple(
        int(piece) if piece.isdigit() else piece.casefold()
        for piece in re.split(r"(\d+)", value)
    )


def _parse_routes(
    rows: Sequence[Sequence[Any]],
    settings: Mapping[str, Any],
) -> list[dict[str, Any]]:
    header_row, headers = _find_header(
        rows,
        "Approved Routes",
        ROUTE_HEADERS,
        10,
    )
    routes: list[dict[str, Any]] = []
    part_operations: set[str] = set()
    sequences_by_part: dict[str, set[int]] = {}

    for row_index in range(header_row + 1, len(rows)):
        row = rows[row_index]
        part = _text(_cell(row, headers, "Part"))
        part_operation = _text(_cell(row, headers, "Part + Operation"))
        if not part and not part_operation:
            continue
        if not part or not part_operation:
            raise WorkbookValidationError(
                f"Approved Routes row {row_index + 1} is missing Part or "
                "Part + Operation."
            )
        if len(routes) >= MAX_ROUTES:
            raise WorkbookValidationError(
                f"Approved Routes exceeds the {MAX_ROUTES:,}-row limit."
            )

        sequence = _require_integer(
            _cell(row, headers, "Numeric Seq"),
            f"Approved Routes row {row_index + 1} Numeric Seq",
            1,
            9_999,
        )
        sequence_override = _optional_integer(
            _cell(row, headers, "Seq Override"),
            f"Approved Routes row {row_index + 1} Seq Override",
            1,
            9_999,
        )
        effective_sequence = (
            sequence if sequence_override is None else sequence_override
        )
        operation = _require_integer(
            _cell(row, headers, "Operation"),
            f"Approved Routes row {row_index + 1} Operation",
            1,
            999_999,
        )
        expected_part_operation = f"{part}-{operation}"
        if part_operation != expected_part_operation:
            raise WorkbookValidationError(
                f"Approved Routes row {row_index + 1} Part + Operation must "
                f"equal {expected_part_operation}."
            )
        if part_operation in part_operations:
            raise WorkbookValidationError(
                f"Approved Routes contains duplicate {part_operation}."
            )
        part_operations.add(part_operation)
        part_sequences = sequences_by_part.setdefault(part, set())
        if effective_sequence in part_sequences:
            raise WorkbookValidationError(
                f"{part} contains duplicate Effective Seq {effective_sequence}."
            )
        part_sequences.add(effective_sequence)

        work_center = _text(_cell(row, headers, "Work Center"))
        if not work_center:
            raise WorkbookValidationError(
                f"Approved Routes row {row_index + 1} needs a Work Center."
            )
        suggested_skill = _text(_cell(row, headers, "Suggested Skill"))
        skill_override = _text(_cell(row, headers, "Skill Override")) or None
        skill = skill_override or suggested_skill
        if skill not in SKILL_NAMES and _normalized(skill) != "unmapped":
            raise WorkbookValidationError(
                f'Approved Routes row {row_index + 1} uses unknown '
                f'qualification "{skill}".'
            )

        exact_p75_hours_per_piece = _optional_number(
            _cell(row, headers, "Exact P75 Hr/Pc"),
            f"Approved Routes row {row_index + 1} Exact P75 Hr/Pc",
        )
        p75_override_hours_per_piece = _optional_number(
            _cell(row, headers, "P75 Override"),
            f"Approved Routes row {row_index + 1} P75 Override",
        )
        p75_hours_per_piece = (
            exact_p75_hours_per_piece
            if p75_override_hours_per_piece is None
            else p75_override_hours_per_piece
        )
        if p75_hours_per_piece is not None and float(p75_hours_per_piece) <= 0:
            raise WorkbookValidationError(
                f"Approved Routes row {row_index + 1} P75 time must be "
                "greater than zero."
            )
        move_minutes_override = _optional_number(
            _cell(row, headers, "Move Min Override"),
            f"Approved Routes row {row_index + 1} Move Min Override",
        )
        if (
            move_minutes_override is not None
            and not 0 <= float(move_minutes_override) <= 240
        ):
            raise WorkbookValidationError(
                f"Approved Routes row {row_index + 1} Move Min Override must "
                "be 0–240."
            )
        move_minutes = (
            0
            if effective_sequence == 1
            else (
                settings["defaultMoveMinutes"]
                if move_minutes_override is None
                else move_minutes_override
            )
        )
        samples_value = _optional_number(
            _cell(row, headers, "Samples"),
            f"Approved Routes row {row_index + 1} Samples",
        )
        samples = 0 if samples_value is None else samples_value
        wc_source = _text(_cell(row, headers, "WC Source"))
        notes = _text(_cell(row, headers, "Notes")) or None
        flags: list[str] = []
        if p75_hours_per_piece is None:
            flags.append("Missing P75")
        if float(samples) > 0 and float(samples) < 5:
            flags.append("Low sample")
        if wc_source and wc_source != "Exact routing":
            flags.append("Review work-center source")

        routes.append(
            {
                "id": f"{part}|{effective_sequence}",
                "part": part,
                "sequence": sequence,
                "sequenceOverride": sequence_override,
                "effectiveSequence": effective_sequence,
                "partOperation": part_operation,
                "operation": operation,
                "workCenter": work_center,
                "description": _text(
                    _cell(row, headers, "Routing Description")
                ),
                "orangeBucket": _text(_cell(row, headers, "Orange Bucket")),
                "suggestedSkill": suggested_skill,
                "skillOverride": skill_override,
                "skill": skill,
                "exactP75HoursPerPiece": exact_p75_hours_per_piece,
                "p75OverrideHoursPerPiece": p75_override_hours_per_piece,
                "p75HoursPerPiece": p75_hours_per_piece,
                "missingP75": p75_hours_per_piece is None,
                "moveMinutesOverride": move_minutes_override,
                "moveMinutes": move_minutes,
                "samples": samples,
                "historyStatus": _text(
                    _cell(row, headers, "History Status")
                ),
                "notes": notes,
                "flags": flags,
            }
        )

    if not routes:
        raise WorkbookValidationError(
            "Approved Routes does not contain any usable routes."
        )
    routes.sort(
        key=lambda route: (
            _natural_key(route["part"]),
            route["effectiveSequence"],
        )
    )
    return routes


def _parse_attendance(
    rows: Sequence[Sequence[Any]],
) -> dict[str, dict[str, dict[str, Any]]]:
    header_row, headers = _find_header(
        rows,
        "Attendance",
        ATTENDANCE_HEADERS,
        10,
    )
    by_id: dict[str, dict[str, Any]] = {}
    by_name: dict[str, dict[str, Any]] = {}

    for row_index in range(header_row + 1, len(rows)):
        row = rows[row_index]
        employee_id = _text(_cell(row, headers, "Employee ID"))
        name = _text(_cell(row, headers, "Employee Name"))
        if not employee_id and not name:
            continue
        label = name or employee_id or f"row {row_index + 1}"
        hours_override = _optional_number(
            _cell(row, headers, "Hours Override"),
            f"Attendance row {row_index + 1} Hours Override",
        )
        if hours_override is not None and not 0 <= float(hours_override) <= 24:
            raise WorkbookValidationError(
                f"Attendance Hours Override for {label} must be 0–24."
            )
        record = {
            "employeeId": employee_id,
            "name": name,
            "scheduled": _yes_no(
                _cell(row, headers, "Scheduled Today?"),
                True,
                f"Attendance row {row_index + 1} Scheduled Today?",
            ),
            "present": _yes_no(
                _cell(row, headers, "Present Today?"),
                True,
                f"Attendance row {row_index + 1} Present Today?",
            ),
            "shiftToday": _optional_integer(
                _cell(row, headers, "Shift Today"),
                f"Attendance row {row_index + 1} Shift Today",
                1,
                999,
            ),
            "hoursOverride": hours_override,
            "sourceHours": _text(_cell(row, headers, "Source Hours")),
            "sourceDays": _text(_cell(row, headers, "Source Days")),
            "dataQualityNote": (
                _text(_cell(row, headers, "Data Quality Note")) or None
            ),
        }
        if employee_id:
            by_id[employee_id] = record
        if name:
            by_name[_normalized(name)] = record
    return {"byId": by_id, "byName": by_name}


def _parse_operators(
    rows: Sequence[Sequence[Any]],
    attendance: Mapping[str, Mapping[str, dict[str, Any]]],
    settings: Mapping[str, Any],
    warnings: list[str],
) -> list[dict[str, Any]]:
    header_row, headers = _find_header(
        rows,
        "Operator Skills",
        OPERATOR_HEADERS,
        10,
    )
    operators: list[dict[str, Any]] = []
    ids: set[str] = set()
    names: set[str] = set()
    matched_attendance: set[int] = set()

    for row_index in range(header_row + 1, len(rows)):
        row = rows[row_index]
        employee_id = _text(_cell(row, headers, "Employee ID"))
        name = _text(_cell(row, headers, "Employee Name"))
        if not employee_id and not name:
            continue
        if not name:
            raise WorkbookValidationError(
                f"Operator Skills row {row_index + 1} needs an Employee Name."
            )
        if len(operators) >= MAX_OPERATORS:
            raise WorkbookValidationError(
                f"Operator Skills exceeds the {MAX_OPERATORS:,}-row limit."
            )
        name_key = _normalized(name)
        if name_key in names:
            raise WorkbookValidationError(
                f'Operator Skills contains duplicate name "{name}".'
            )
        names.add(name_key)
        if employee_id:
            if employee_id in ids:
                raise WorkbookValidationError(
                    f"Operator Skills contains duplicate Employee ID {employee_id}."
                )
            ids.add(employee_id)

        source_shift = _optional_integer(
            _cell(row, headers, "Source Shift"),
            f"Operator Skills row {row_index + 1} Source Shift",
            1,
            999,
        )
        attendance_record = (
            attendance["byId"].get(employee_id) if employee_id else None
        ) or attendance["byName"].get(name_key)
        if attendance_record is not None:
            matched_attendance.add(id(attendance_record))
        else:
            warnings.append(
                f"{name} has no matching Attendance row; "
                "scheduled/present defaults were used."
            )

        shift_candidate = (
            attendance_record["shiftToday"]
            if attendance_record is not None
            and attendance_record["shiftToday"] is not None
            else source_shift
        )
        shift_ids = {shift["id"] for shift in settings["shifts"]}
        default_shift = (
            shift_candidate
            if shift_candidate in shift_ids
            else settings["shifts"][0]["id"]
        )
        shift = next(
            item for item in settings["shifts"] if item["id"] == default_shift
        )

        skills: dict[str, int] = {}
        for skill in SKILL_NAMES:
            raw_skill = _optional_number(
                _cell(row, headers, skill),
                f"Operator Skills row {row_index + 1} {skill}",
            )
            value = 0 if raw_skill is None else raw_skill
            if (
                not float(value).is_integer()
                or float(value) < 0
                or float(value) > 3
            ):
                raise WorkbookValidationError(
                    f"Operator Skills row {row_index + 1} {skill} must be a "
                    "whole number from 0 to 3."
                )
            skills[skill] = int(value)

        hours_override = (
            attendance_record["hoursOverride"]
            if attendance_record is not None
            else None
        )
        operators.append(
            {
                "schedulerKey": (
                    f"id:{employee_id}"
                    if employee_id
                    else f"name:{name_key}"
                ),
                "employeeId": employee_id or None,
                "name": name,
                "sourceShift": source_shift,
                "defaultShift": default_shift,
                "defaultScheduled": (
                    attendance_record["scheduled"]
                    if attendance_record is not None
                    else True
                ),
                "defaultPresent": (
                    attendance_record["present"]
                    if attendance_record is not None
                    else True
                ),
                "defaultHoursOverride": hours_override,
                "defaultAvailableHours": (
                    shift["availableHours"]
                    if hours_override is None
                    else hours_override
                ),
                "sourceHours": (
                    attendance_record["sourceHours"]
                    if attendance_record is not None
                    else ""
                ),
                "sourceDays": (
                    attendance_record["sourceDays"]
                    if attendance_record is not None
                    else ""
                ),
                "skills": skills,
                "dataQualityNote": (
                    _text(_cell(row, headers, "Data Quality Note"))
                    or (
                        attendance_record["dataQualityNote"]
                        if attendance_record is not None
                        else None
                    )
                    or None
                ),
            }
        )

    if not operators:
        raise WorkbookValidationError(
            "Operator Skills does not contain any employees."
        )
    for record in attendance["byId"].values():
        if id(record) not in matched_attendance:
            warnings.append(
                f"{record['name'] or record['employeeId']} appears in "
                "Attendance but not Operator Skills and was skipped."
            )
    return operators


def _assert_acyclic(
    graph: Mapping[str, Sequence[str]],
) -> None:
    state: dict[str, int] = {}
    for start in graph:
        if state.get(start, 0) == 2:
            continue
        state[start] = 1
        stack: list[tuple[str, Iterable[str]]] = [
            (start, iter(graph.get(start, ())))
        ]
        while stack:
            part, children = stack[-1]
            try:
                child = next(children)
            except StopIteration:
                state[part] = 2
                stack.pop()
                continue
            child_state = state.get(child, 0)
            if child_state == 1:
                raise WorkbookValidationError(
                    f"Parent-Child contains a dependency cycle at {child}."
                )
            if child_state == 0:
                state[child] = 1
                stack.append((child, iter(graph.get(child, ()))))


def _parse_dependencies(
    rows: Sequence[Sequence[Any]],
    route_parts: set[str],
) -> list[dict[str, Any]]:
    header_row, headers = _find_header(
        rows,
        "Parent-Child",
        DEPENDENCY_HEADERS,
        10,
    )
    pairs: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for row_index in range(header_row + 1, len(rows)):
        row = rows[row_index]
        parent = _text(_cell(row, headers, "Parent Part"))
        child = _text(_cell(row, headers, "Child Part"))
        if not parent and not child:
            continue
        if not parent or not child:
            raise WorkbookValidationError(
                f"Parent-Child row {row_index + 1} needs both Parent Part "
                "and Child Part."
            )
        if len(pairs) >= MAX_DEPENDENCIES:
            raise WorkbookValidationError(
                f"Parent-Child exceeds the {MAX_DEPENDENCIES:,}-row limit."
            )
        if parent == child:
            raise WorkbookValidationError(
                f"{parent} cannot be its own child part."
            )
        if parent not in route_parts or child not in route_parts:
            raise WorkbookValidationError(
                f"Parent-Child row {row_index + 1} references a part without "
                "an approved route."
            )
        key = (parent, child)
        if key in seen:
            raise WorkbookValidationError(
                f"Parent-Child contains duplicate {parent} → {child}."
            )
        seen.add(key)
        pairs.append(
            {
                "id": len(pairs) + 1,
                "parent": parent,
                "child": child,
            }
        )

    graph: dict[str, list[str]] = {}
    for pair in pairs:
        graph.setdefault(pair["parent"], []).append(pair["child"])
    _assert_acyclic(graph)
    return pairs


def _parse_loaded_workbook(
    workbook: Any,
    file_name: str | None,
) -> dict[str, Any]:
    if workbook is None or not hasattr(workbook, "sheetnames"):
        raise WorkbookValidationError(
            "The selected file is not a readable Excel workbook."
        )

    rows_by_sheet: dict[str, list[tuple[Any, ...]]] = {}
    for sheet_name, title in REQUIRED_SHEETS.items():
        rows = _rows_for(workbook, sheet_name)
        _verify_title(rows, sheet_name, title)
        rows_by_sheet[sheet_name] = rows

    warnings: list[str] = []
    settings = _parse_settings(rows_by_sheet["Settings"], warnings)
    routes = _parse_routes(rows_by_sheet["Approved Routes"], settings)
    attendance = _parse_attendance(rows_by_sheet["Attendance"])
    operators = _parse_operators(
        rows_by_sheet["Operator Skills"],
        attendance,
        settings,
        warnings,
    )
    route_parts = {route["part"] for route in routes}
    parent_child = _parse_dependencies(
        rows_by_sheet["Parent-Child"],
        route_parts,
    )

    imported_at = (
        datetime.now(timezone.utc)
        .isoformat(timespec="milliseconds")
        .replace("+00:00", "Z")
    )
    return {
        "data": {
            "schemaVersion": 2,
            "source": {
                "workbook": file_name or "Imported scheduler workbook.xlsx",
                "importedAt": imported_at,
                "sheets": list(REQUIRED_SHEETS),
            },
            "settings": settings,
            "skillNames": list(SKILL_NAMES),
            "operators": operators,
            "routes": routes,
            "parentChild": parent_child,
        },
        "warnings": warnings,
        "counts": {
            "routes": len(routes),
            "completeParts": len(route_parts),
            "operators": len(operators),
            "relationships": len(parent_child),
            "shifts": len(settings["shifts"]),
        },
    }


def parse_workbook(
    workbook: Any,
    file_name: str | None = None,
) -> dict[str, Any]:
    """Parse an already-loaded openpyxl workbook.

    Bytes and file paths are also accepted as a convenience; application code
    should normally call :func:`parse_workbook_bytes` for uploaded files.
    """

    if isinstance(workbook, (bytes, bytearray, memoryview)):
        return parse_workbook_bytes(bytes(workbook), file_name)
    if isinstance(workbook, (str, Path)):
        return parse_workbook_file(workbook)
    return _parse_loaded_workbook(workbook, file_name)


def parse_workbook_bytes(
    data: bytes,
    file_name: str | None = None,
) -> dict[str, Any]:
    """Validate and parse an uploaded ``.xlsx`` workbook into schema v2."""

    if not isinstance(data, (bytes, bytearray, memoryview)):
        raise WorkbookValidationError(
            "The selected file is not a readable Excel workbook."
        )
    raw = bytes(data)
    effective_name = file_name or "Imported scheduler workbook.xlsx"
    validate_workbook_file(effective_name, len(raw))
    try:
        workbook = load_workbook(
            BytesIO(raw),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise WorkbookValidationError(
            "The selected file is not a readable Excel workbook."
        ) from exc
    try:
        return _parse_loaded_workbook(workbook, effective_name)
    finally:
        workbook.close()


def parse_workbook_file(path: str | Path) -> dict[str, Any]:
    """Load and parse a workbook from a local file path."""

    workbook_path = Path(path)
    try:
        size_bytes = workbook_path.stat().st_size
    except OSError as exc:
        raise WorkbookValidationError(
            "The selected file could not be read."
        ) from exc
    validate_workbook_file(workbook_path.name, size_bytes)
    try:
        data = workbook_path.read_bytes()
    except OSError as exc:
        raise WorkbookValidationError(
            "The selected file could not be read."
        ) from exc
    return parse_workbook_bytes(data, workbook_path.name)


__all__ = [
    "ATTENDANCE_HEADERS",
    "DEPENDENCY_HEADERS",
    "MAX_DEPENDENCIES",
    "MAX_OPERATORS",
    "MAX_ROUTES",
    "MAX_WORKBOOK_BYTES",
    "OPERATOR_HEADERS",
    "REQUIRED_HEADERS",
    "REQUIRED_SHEETS",
    "ROUTE_HEADERS",
    "SETTINGS_HEADERS",
    "SHIFT_HEADERS",
    "SKILL_NAMES",
    "WorkbookValidationError",
    "parse_workbook",
    "parse_workbook_bytes",
    "parse_workbook_file",
    "validate_workbook_file",
    "validate_workbook_upload",
]
