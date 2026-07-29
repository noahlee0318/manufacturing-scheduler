from __future__ import annotations

import unittest
from datetime import time
from io import BytesIO

from openpyxl import Workbook

from shift_scheduler.workbook import (
    ATTENDANCE_HEADERS,
    DEPENDENCY_HEADERS,
    MAX_WORKBOOK_BYTES,
    OPERATOR_HEADERS,
    REQUIRED_SHEETS,
    ROUTE_HEADERS,
    SCHEDULING_RULES,
    SETTINGS_HEADERS,
    SHIFT_HEADERS,
    SKILL_NAMES,
    WorkbookValidationError,
    parse_workbook,
    parse_workbook_bytes,
    validate_workbook_file,
)


def _append_table(
    worksheet,
    title: str,
    headers: tuple[str, ...],
    rows: list[list[object]],
) -> None:
    worksheet["A1"] = title
    worksheet.append([])
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(row)


def _route_row(
    part: str,
    sequence: int,
    operation: int,
    *,
    skill: str = "Tube",
    exact_p75: float | None = 0.5,
) -> list[object]:
    values = {
        "Part": part,
        "Numeric Seq": sequence,
        "Seq Override": None,
        "Part + Operation": f"{part}-{operation}",
        "Operation": operation,
        "Work Center": f"WC-{sequence}",
        "Routing Description": "Demonstration operation",
        "Orange Bucket": "Demonstration bucket",
        "Suggested Skill": skill,
        "Skill Override": None,
        "Exact P75 Hr/Pc": exact_p75,
        "P75 Override": None,
        "Move Min Override": None,
        "Samples": 12,
        "History Status": "Good sample",
        "WC Source": "Exact routing",
        "Orange Cell": "A1",
        "Notes": None,
    }
    return [values[header] for header in ROUTE_HEADERS]


def _operator_skills() -> dict[str, int]:
    return {
        skill: level
        for skill, level in zip(
            SKILL_NAMES,
            (3, 2, 1, 0, 3, 2, 1, 0, 3, 2, 1, 0, 3, 2),
            strict=True,
        )
    }


def _make_workbook() -> Workbook:
    workbook = Workbook()

    routes = workbook.active
    routes.title = "Approved Routes"
    _append_table(
        routes,
        REQUIRED_SHEETS["Approved Routes"],
        ROUTE_HEADERS,
        [
            _route_row("PART-A", 1, 10),
            _route_row("PART-B", 1, 20, skill="Spray Bars"),
        ],
    )

    operators = workbook.create_sheet("Operator Skills")
    skills = _operator_skills()
    operator_values = {
        "Employee ID": "EMP-001",
        "Employee Name": "Employee 1",
        "Source Shift": 1,
        **skills,
        "Data Quality Note": None,
    }
    _append_table(
        operators,
        REQUIRED_SHEETS["Operator Skills"],
        OPERATOR_HEADERS,
        [[operator_values[header] for header in OPERATOR_HEADERS]],
    )

    attendance = workbook.create_sheet("Attendance")
    attendance_values = {
        "Employee ID": "EMP-001",
        "Employee Name": "Employee 1",
        "Scheduled Today?": "Yes",
        "Present Today?": "Yes",
        "Shift Today": 1,
        "Hours Override": None,
        "Source Hours": "05:00-15:30",
        "Source Days": "Mon-Thu",
        "Data Quality Note": None,
    }
    _append_table(
        attendance,
        REQUIRED_SHEETS["Attendance"],
        ATTENDANCE_HEADERS,
        [[attendance_values[header] for header in ATTENDANCE_HEADERS]],
    )

    dependencies = workbook.create_sheet("Parent-Child")
    _append_table(
        dependencies,
        REQUIRED_SHEETS["Parent-Child"],
        DEPENDENCY_HEADERS,
        [],
    )

    settings = workbook.create_sheet("Settings")
    settings["A1"] = REQUIRED_SHEETS["Settings"]
    settings.append([])
    settings.append(list(SETTINGS_HEADERS))
    settings.append(
        ["Minimum solo skill level", 2, "Independent or trainer"]
    )
    settings.append(["Default move minutes", 5, "Travel between route steps"])
    settings.append(["Work-center capacity", 1, "One operation at a time"])
    settings.append([])
    settings.append(list(SHIFT_HEADERS))
    settings.append([1, time(5, 0), time(15, 30), 10.5, "Configured Shift 1"])
    settings.append([2, time(15, 30), time(2, 0), 10.5, "Configured Shift 2"])

    return workbook


def _set_header_value(
    workbook: Workbook,
    sheet_name: str,
    header: str,
    row_number: int,
    value: object,
    *,
    header_row: int = 3,
) -> None:
    worksheet = workbook[sheet_name]
    headers = [
        worksheet.cell(header_row, column).value
        for column in range(1, worksheet.max_column + 1)
    ]
    worksheet.cell(row_number, headers.index(header) + 1).value = value


def _as_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class WorkbookImportTests(unittest.TestCase):
    def test_valid_workbook_parses_to_schema_v2(self) -> None:
        result = parse_workbook_bytes(
            _as_bytes(_make_workbook()),
            "schedule.xlsx",
        )

        self.assertEqual(result["data"]["schemaVersion"], 2)
        self.assertEqual(result["data"]["source"]["workbook"], "schedule.xlsx")
        self.assertEqual(
            result["data"]["source"]["sheets"],
            list(REQUIRED_SHEETS),
        )
        self.assertEqual(
            result["data"]["settings"]["shifts"][1],
            {
                "id": 2,
                "startMinutes": 930,
                "endMinutes": 120,
                "startTime": "15:30",
                "endTime": "02:00",
                "availableHours": 10.5,
                "crossesMidnight": True,
                "source": "Configured Shift 2",
                "placeholder": False,
            },
        )
        self.assertEqual(result["warnings"], [])
        self.assertEqual(
            result["data"]["settings"]["schedulingRules"],
            SCHEDULING_RULES,
        )
        self.assertEqual(
            SCHEDULING_RULES["Batch"],
            "Each whole unit may advance after its own preceding route operation "
            "finishes; it does not wait for the remaining quantity",
        )
        self.assertEqual(
            SCHEDULING_RULES["Dispatch"],
            "Active children before parents; then priority, due time, shorter route "
            "time, stable job order, unit, and route step",
        )
        self.assertEqual(
            result["counts"],
            {
                "routes": 2,
                "completeParts": 2,
                "operators": 1,
                "relationships": 0,
                "shifts": 2,
            },
        )

    def test_missing_sheet_is_rejected(self) -> None:
        workbook = _make_workbook()
        del workbook["Attendance"]

        with self.assertRaisesRegex(
            WorkbookValidationError,
            "Attendance.*missing",
        ):
            parse_workbook(workbook)

    def test_wrong_title_is_rejected(self) -> None:
        workbook = _make_workbook()
        workbook["Operator Skills"]["A1"] = "A different workbook"

        with self.assertRaisesRegex(WorkbookValidationError, "Cell A1"):
            parse_workbook(workbook)

    def test_missing_required_header_is_rejected(self) -> None:
        workbook = _make_workbook()
        _set_header_value(
            workbook,
            "Approved Routes",
            "Notes",
            3,
            "Comments",
        )

        with self.assertRaisesRegex(
            WorkbookValidationError,
            "missing required columns",
        ):
            parse_workbook(workbook)

    def test_duplicate_part_operation_is_rejected(self) -> None:
        workbook = _make_workbook()
        workbook["Approved Routes"].append(_route_row("PART-A", 2, 10))

        with self.assertRaisesRegex(
            WorkbookValidationError,
            "duplicate PART-A-10",
        ):
            parse_workbook(workbook)

    def test_duplicate_effective_sequence_is_rejected(self) -> None:
        workbook = _make_workbook()
        workbook["Approved Routes"].append(_route_row("PART-A", 1, 30))

        with self.assertRaisesRegex(
            WorkbookValidationError,
            "duplicate Effective Seq 1",
        ):
            parse_workbook(workbook)

    def test_unknown_route_skill_is_rejected(self) -> None:
        workbook = _make_workbook()
        _set_header_value(
            workbook,
            "Approved Routes",
            "Suggested Skill",
            4,
            "Mystery Skill",
        )

        with self.assertRaisesRegex(
            WorkbookValidationError,
            "unknown qualification",
        ):
            parse_workbook(workbook)

    def test_missing_p75_is_accepted_and_flagged(self) -> None:
        workbook = _make_workbook()
        _set_header_value(
            workbook,
            "Approved Routes",
            "Exact P75 Hr/Pc",
            4,
            None,
        )

        route = parse_workbook(workbook)["data"]["routes"][0]

        self.assertEqual(route["part"], "PART-A")
        self.assertIsNone(route["p75HoursPerPiece"])
        self.assertIs(route["missingP75"], True)
        self.assertEqual(route["flags"], ["Missing P75"])

    def test_parent_child_cycle_is_rejected(self) -> None:
        workbook = _make_workbook()
        dependencies = workbook["Parent-Child"]
        dependencies.append(["PART-A", "PART-B"])
        dependencies.append(["PART-B", "PART-A"])

        with self.assertRaisesRegex(
            WorkbookValidationError,
            "dependency cycle",
        ):
            parse_workbook(workbook)

    def test_work_center_capacity_must_remain_one(self) -> None:
        workbook = _make_workbook()
        workbook["Settings"]["B6"] = 2

        with self.assertRaisesRegex(
            WorkbookValidationError,
            "capacity to remain 1",
        ):
            parse_workbook(workbook)

    def test_invalid_optional_number_is_rejected(self) -> None:
        cases = [
            (
                "Approved Routes",
                "Seq Override",
                4,
                3,
                "later",
                "Seq Override must be a number or blank",
            ),
            (
                "Approved Routes",
                "Exact P75 Hr/Pc",
                4,
                3,
                "fast",
                "Exact P75 Hr/Pc must be a number or blank",
            ),
            (
                "Approved Routes",
                "Samples",
                4,
                3,
                "many",
                "Samples must be a number or blank",
            ),
            (
                "Operator Skills",
                "Source Shift",
                4,
                3,
                "first",
                "Source Shift must be a number or blank",
            ),
            (
                "Operator Skills",
                "Tube",
                4,
                3,
                "expert",
                "Tube must be a number or blank",
            ),
            (
                "Attendance",
                "Shift Today",
                4,
                3,
                "first",
                "Shift Today must be a number or blank",
            ),
            (
                "Attendance",
                "Hours Override",
                4,
                3,
                "full",
                "Hours Override must be a number or blank",
            ),
            (
                "Settings",
                "Available Hours",
                9,
                8,
                "ten",
                "Available Hours must be a number or blank",
            ),
        ]
        for (
            sheet_name,
            header,
            row_number,
            header_row,
            value,
            message,
        ) in cases:
            with self.subTest(sheet=sheet_name, header=header):
                workbook = _make_workbook()
                _set_header_value(
                    workbook,
                    sheet_name,
                    header,
                    row_number,
                    value,
                    header_row=header_row,
                )
                with self.assertRaisesRegex(
                    WorkbookValidationError,
                    message,
                ):
                    parse_workbook(workbook)

    def test_invalid_optional_yes_no_is_rejected(self) -> None:
        workbook = _make_workbook()
        _set_header_value(
            workbook,
            "Attendance",
            "Present Today?",
            4,
            "Maybe",
        )

        with self.assertRaisesRegex(
            WorkbookValidationError,
            "Yes, No, or blank",
        ):
            parse_workbook(workbook)

    def test_all_fourteen_skill_levels_are_preserved_exactly(self) -> None:
        result = parse_workbook(_make_workbook())
        operator = result["data"]["operators"][0]

        self.assertEqual(result["data"]["skillNames"], list(SKILL_NAMES))
        self.assertEqual(len(operator["skills"]), 14)
        self.assertEqual(operator["skills"], _operator_skills())

    def test_upload_validation_requires_xlsx_and_enforces_ten_mib(self) -> None:
        validate_workbook_file("schedule.XLSX", MAX_WORKBOOK_BYTES)

        with self.assertRaisesRegex(WorkbookValidationError, r"\.xlsx"):
            validate_workbook_file("schedule.xls", 100)
        with self.assertRaisesRegex(WorkbookValidationError, "10 MiB"):
            validate_workbook_file(
                "schedule.xlsx",
                MAX_WORKBOOK_BYTES + 1,
            )


if __name__ == "__main__":
    unittest.main()
